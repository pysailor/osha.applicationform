from collective.lead.interfaces import IDatabase
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from DateTime import DateTime
from logging import getLogger
from openpyxl.workbook import Workbook
from osha.applicationform import _
from osha.applicationform.config import COUNTRIES
from osha.applicationform.config import NATIONALITIES
from osha.applicationform.config import OSHA_HR_EMAIL
from osha.applicationform.config import PFG_FILE_UPLOAD_PREFIX
from plone import api
from plone.i18n.locales.languages import contentlanguages
from Products.CMFCore.utils import getToolByName
from Products.Five.browser import BrowserView
from Products.PFGDataGrid.vocabulary import SimpleDynamicVocabulary
from slc.rdbploneformgenadapter.utils import cleanString
from StringIO import StringIO
from smtplib import SMTPRecipientsRefused
from zipfile import ZipFile
from zipfile import ZipInfo
from zope.component import getUtility
from zope.component import ComponentLookupError


logger = getLogger('osha.applicationform.views')


SEND_DATA_EMAIL_MESSAGE = u"""Hello,

we're sending you a file with saved user submissions.

Best regards,
{0}
"""


class VacanciesView(BrowserView):
    """Helper view to get the available job vacancies."""

    def __call__(self):
        catalog = getToolByName(self.context, 'portal_catalog')
        results = catalog(
            portal_type='PublicJobVacancy',
            review_state='published',
            sort_on='effective',
        )
        # return job vacancies that are not past deadline
        return [(item.getId, item.Title) for item in results
                if DateTime(datetime.now()) < item.getObject().deadline]


class CountriesView(BrowserView):
    """Helper view to get a list of countries."""

    def __call__(self):
        return [(item.lower(), item) for item in COUNTRIES]


class LanguagesView(BrowserView):
    """Helper view to get a list of languages."""

    def __call__(self):
        languages = sorted(
            [lang[1] for lang in contentlanguages.getLanguageListing()])
        return SimpleDynamicVocabulary(languages)


class NationalitiesView(BrowserView):
    """Helper view to get a list of nationalities."""

    def __call__(self):
        return [(item.lower(), item) for item in NATIONALITIES]


class PFGSaveDataAdapterWithFileUploadView(BrowserView):
    """View for displaying saved data (with links to uploaded files) and
    option to export the results.

    INFO: Template is a combination of fg_savedata_view_p3.pt and
    fg_savedata_tabview_p3.pt from Products.PloneFormGen.
    """

    def __call__(self):
        return self.index()

    def get_file_url(self, field_name, submission_id):
        """Return url to the uploaded file.

        :param field_name: name of the field that we want to fetch the file
            url for
        :submission_id: id of the submission
        """

        # We need the file prefix to identify the file in the results row.
        submission_uuid = submission_id.split(PFG_FILE_UPLOAD_PREFIX)[1]
        uploads = self.context.getParentNode().get('uploads', None)

        if uploads:
            return "{0}/{1}/{2}/view".format(
                uploads.absolute_url(), submission_uuid, field_name)
        else:
            return None


class ExportSavedDataWithFiles(BrowserView):
    """View for exporting saved data and uploaded files."""

    def __call__(self):
        """Download the data in a zip package."""
        filename = self.context.id
        self.context.REQUEST.response.setHeader(
            "Content-Type",
            "application/zip"
        )
        self.context.REQUEST.response.setHeader(
            'Content-Disposition',
            "attachment; filename=%s.zip" % filename
        )
        return self.context.get_data_zipped()


class SendSavedDataWithFiles(BrowserView):
    """View for sending saved data and uploaded files via email."""

    def __call__(self):
        portal = api.portal.get()
        site_properties = api.portal.get_tool(
            'portal_properties').site_properties
        email_charset = site_properties.getProperty('email_charset', 'utf-8')
        sender = portal.getProperty('email_from_address', '')

        # create the container (outer) email message.
        msg = MIMEMultipart()
        msg['Subject'] = _(u'HR applications data')
        msg['From'] = sender
        msg['To'] = OSHA_HR_EMAIL

        text = SEND_DATA_EMAIL_MESSAGE.format(
            portal.getProperty('title')).encode(email_charset, 'replace')

        # add body text
        body = MIMEText(
            text,
            _subtype='plain',
            _charset=email_charset
        )
        msg.attach(body)

        # add zip file
        attachment = MIMEBase('application', 'zip')
        attachment.add_header(
            'Content-Disposition',
            'attachment',
            filename=self.context.getParentNode().id + '.zip'
        )
        attachment.set_payload(self.get_data_zipped())
        encoders.encode_base64(attachment)
        msg.attach(attachment)

        mailhost = api.portal.get_tool('MailHost')

        # send the message
        try:
            mailhost.send(msg, immediate=True)
        except SMTPRecipientsRefused:
            logger.error('Error sending email to {0}'.format(OSHA_HR_EMAIL))
            raise SMTPRecipientsRefused('Recipient address rejected by server')
        logger.info('Data sent to {0}'.format(OSHA_HR_EMAIL))
        return 'Data sent to {0}'.format(OSHA_HR_EMAIL)

    def write_to_zip(self, zipfile):
        """Write form data to excel file and zip it together with uploaded
        files.

        File columns contain links to uploaded files. File structure of the
        exported excel file and accompanying files will be like this:

        form_id.xlsx
        files
        |
        |--id_some_file1.txt
        |--id_some_file2.pdf
        ...

        There are a couple of assumptions about the rdb structure. Please see
        slc.rdbploneformgenadapter.content.content.py

        :param zipfile: ZipFile instance that will hold the data
        """
        form = self.context.getParentNode()
        form_table_name = cleanString(form.id)
        db_utility_name = self.context.db_utility_name

        try:
            db = getUtility(IDatabase, db_utility_name)
        except ComponentLookupError:
            api.portal.show_message(
                message="""There is a problem with database configuration,
                contact site administrator""",
                request=self.request,
                type='error'
            )
        connection = db.connection.engine.connect()

        grid_fields = []
        file_fields = []
        for key in form.keys():
            if(form[key].portal_type == 'FormDataGridField'):
                grid_fields.append(key)
            elif(form[key].portal_type == 'FormFileField'):
                file_fields.append(key)
        form_data = connection.execute(
            "SELECT * FROM {0}".format(form_table_name)
        )

        # Extending form_keys with grid and file fields, so we can
        # write them to excel easier inside the loop
        form_keys = form_data.keys()
        all_form_keys = form_keys + grid_fields + file_fields

        book = Workbook()
        sheet = book.worksheets[0]
        sheet.default_column_dimension.auto_size = True
        sheet.title = form_table_name
        date_time = datetime.now().timetuple()[:6]

        # write column names
        for column_count, key in enumerate(all_form_keys):
            sheet.cell(row=0,
                       column=column_count).value = all_form_keys[column_count]
        # write data
        for row_count, table_row in enumerate(form_data.fetchall()):
            # write basic form fields
            for form_column_count, key in enumerate(form_keys):
                cell = sheet.cell(row=row_count+1,
                                  column=form_column_count)
                cell.value = table_row[key]
            count = 1

            # write grid form fields
            for field in grid_fields:
                grid_table_name = form_table_name + "_" + cleanString(field)
                grid_field_data = connection.execute(
                    'SELECT * FROM {0} WHERE {1}_id={2}'.format(
                    grid_table_name, form_table_name, table_row[0])
                )
                # hack to get all grid rows in same cell
                grid_field_string = ""
                keys = grid_field_data.keys()
                keys.remove(form_table_name + '_id')
                grid_field_data_all = grid_field_data.fetchall()
                for row in grid_field_data_all:
                    for key in keys:
                        grid_field_string += "{0}: {1}, ".format(key, row[key])
                    grid_field_string = grid_field_string[:-2] + '\n'
                if grid_field_string:
                    cell = sheet.cell(
                        row=row_count+1,
                        column=form_column_count + count
                    )
                    cell.style.alignment.wrap_text = True
                    cell.value = grid_field_string[:-1]
                # set width on longest row
                sheet.column_dimensions[cell.column].width = max(
                    [len(grid_row) for
                        grid_row in grid_field_string.split("\n")]
                )
                count += 1

            # write file form fields
            for file_count, field in enumerate(file_fields):
                file_table_name = form_table_name + "_" + cleanString(field)
                file_field_data = connection.execute(
                    'SELECT * FROM {0} WHERE {1}_id={2}'.format(
                    file_table_name, form_table_name, table_row[0]))
                filename = str(table_row[0]) + "_" + str(file_count)
                for row in file_field_data.fetchall():
                    file_path = 'files/{0}.{1}'.format(filename, row["type"])
                    zip_info = ZipInfo(file_path, date_time)
                    # setting permissions manually to make it work on Linux
                    zip_info.external_attr = 0666 << 16L
                    zipfile.writestr(zip_info, row["data"])
                    cell = sheet.cell(
                        row=row_count + 1,
                        column=form_column_count + count
                    )
                    cell.value = "View file"
                    cell.hyperlink = file_path
                    count += 1

        output_tmp = StringIO()
        book.save(output_tmp)

        zip_info = ZipInfo('{0}.xlsx'.format(form_table_name), date_time)
        # setting permissions manually to make it work on Linux
        zip_info.external_attr = 0666 << 16L
        zipfile.writestr(zip_info, output_tmp.getvalue())

    def get_data_zipped(self):
        """Return the data in a zip package.

        :returns: ZIP file with saved data in excel format and uploaded files
        :rtype: StringIO binary stream
        """

        output = StringIO()
        zipfile = ZipFile(output, mode='w')

        try:
            self.write_to_zip(zipfile)
        finally:
            zipfile.close()

        return output.getvalue()
